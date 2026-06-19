"""Exam configuration importer (V3) — from a manual LearnWorlds UI export.

The assessment configuration (question text, options, CONFIGURED CORRECT ANSWERS,
feedback) is NOT exposed by the LearnWorlds API (confirmed: the whole Assessments
API section is responses + form responses + review only). It must be exported
manually from the UI as an XLSX with a "Questions" sheet.

This module reads that XLSX and normalizes it into exam_config_as_is.csv.
It does NOT call the API, does NOT validate answers, and does NOT compare against
submissions — it only turns the manual export into a clean, audit-friendly table.

Known export quirk (handled): the layout is POSITIONAL, not semantic. For questions
with more than 5 options (e.g. fill-in with many accepted variants), the option list
overflows past `Answer5` and REUSES the `CorrectExplanation` / `IncorrectExplanation`
columns (and further unlabeled columns) as answer slots. We detect this by the max
referenced answer index and only treat col9/col10 as feedback when they are not
consumed as answers.
"""

from __future__ import annotations

import re
from pathlib import Path

from .config import ExtractorError

SHEET_NAME = "Questions"

# Output columns. Fields the manual export does NOT contain are emitted blank and
# documented (assessment_id/course_id/unit_uid/blockId/configured_score/settings_raw
# come from the API/UI context, not from this sheet).
EXAM_CONFIG_COLUMNS = [
    "assessment_id",                  # from --assessment-id (else blank)
    "assessment_title",               # from --assessment-title / filename
    "course_id",                      # from --course-id (else blank)
    "unit_uid",                       # not in export -> blank
    "blockId",                        # not in export -> blank
    "group",                          # Group
    "question_number",                # 1-based order in the sheet
    "blockType",                      # Type (raw: TTF/TMC/TMCMA/TST/...)
    "description",                    # Question
    "options_raw",                    # JSON array of all option cells (positional)
    "configured_correct_answer_raw",  # CorrectAns, verbatim
    "configured_accepted_answers_raw",# JSON array of option texts at correct indices
    "configured_score",               # not in export -> blank
    "feedback_correct",               # CorrectExplanation (when not an answer slot)
    "feedback_incorrect",             # IncorrectExplanation (when not an answer slot)
    "settings_raw",                   # not in export -> blank
    "row_raw",                        # full raw row (JSON) — loss-less safety net
    "source",                         # manual_ui_export:<filename>
    "extraction_timestamp",
]

_INT_RE = re.compile(r"\d+")


def _parse_correct_indices(value) -> list[int]:
    """Parse a CorrectAns cell ('1' / '1, 3, 4' / '1, 2, ... 30') into ints."""
    if value in (None, ""):
        return []
    return [int(m) for m in _INT_RE.findall(str(value))]


def _nonempty(value) -> bool:
    return value not in (None, "") and str(value).strip() != ""


def title_from_filename(path: Path) -> str:
    """Derive a readable assessment title from the export filename.

    e.g. '2026-06-19_teste-de-avaliacao-uc-mercados.xlsx' -> the part after the date.
    """
    stem = path.stem
    # strip a leading YYYY-MM-DD_ if present
    stem = re.sub(r"^\d{4}-\d{2}-\d{2}[_-]", "", stem)
    return stem


def parse_exam_config_xlsx(
    xlsx_path: Path,
    *,
    assessment_id: str = "",
    assessment_title: str = "",
    course_id: str = "",
    extraction_timestamp: str = "",
) -> tuple[list[dict], list[dict], dict]:
    """Parse a UI-exported assessment XLSX.

    Returns (rows, raw_rows, stats):
        rows     : normalized exam_config rows (EXAM_CONFIG_COLUMNS)
        raw_rows : faithful header->value dicts (audit copy, before normalization)
        stats    : counts / warnings for the report
    """
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise ExtractorError(
            "Missing dependency 'openpyxl'. Install requirements first:\n"
            "    pip install -r requirements.txt"
        ) from exc

    if not xlsx_path.exists():
        raise ExtractorError(f"Exam-config XLSX not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.worksheets[0]
    sheet_rows = list(ws.iter_rows(values_only=True))
    if not sheet_rows:
        raise ExtractorError(f"Sheet '{ws.title}' in {xlsx_path.name} is empty.")

    header = sheet_rows[0]
    # Column lookup by header name (first occurrence wins).
    name_to_idx: dict[str, int] = {}
    for i, h in enumerate(header):
        if h not in (None, "") and h not in name_to_idx:
            name_to_idx[str(h)] = i

    required = ["Type", "Question", "CorrectAns", "Answer1"]
    missing = [c for c in required if c not in name_to_idx]
    if missing:
        raise ExtractorError(
            f"Export is missing expected column(s): {', '.join(missing)}. "
            f"Found headers: {[h for h in header if h]}"
        )

    answer_start = name_to_idx["Answer1"]
    answer_cols = [
        name_to_idx[f"Answer{n}"]
        for n in range(1, 6)
        if f"Answer{n}" in name_to_idx
    ]
    correct_expl_idx = name_to_idx.get("CorrectExplanation")
    incorrect_expl_idx = name_to_idx.get("IncorrectExplanation")
    group_idx = name_to_idx.get("Group")

    title = assessment_title or title_from_filename(xlsx_path)
    source = f"manual_ui_export:{xlsx_path.name}"

    def cell(row, i):
        return row[i] if i is not None and i < len(row) else None

    rows: list[dict] = []
    raw_rows: list[dict] = []
    warnings: list[str] = []
    type_counts: dict[str, int] = {}
    overflow_count = 0

    for n, row in enumerate(sheet_rows[1:], start=1):
        # Skip fully empty rows.
        if not any(_nonempty(c) for c in row):
            continue

        # Faithful raw copy (header name, or colN for unlabeled columns).
        raw = {}
        for i, val in enumerate(row):
            key = str(header[i]) if i < len(header) and header[i] else f"col{i}"
            raw[key] = val
        raw_rows.append(raw)

        block_type = cell(row, name_to_idx["Type"])
        type_counts[str(block_type)] = type_counts.get(str(block_type), 0) + 1
        correct_raw = cell(row, name_to_idx["CorrectAns"])
        correct_indices = _parse_correct_indices(correct_raw)
        max_idx = max(correct_indices) if correct_indices else 0

        # Detect positional overflow: >5 referenced options means the option list
        # has consumed the feedback columns (and beyond) as answer slots.
        overflow = max_idx > 5
        if overflow:
            overflow_count += 1
            options = [cell(row, answer_start + (k - 1)) for k in range(1, max_idx + 1)]
            feedback_correct = None
            feedback_incorrect = None
        else:
            options = [cell(row, c) for c in answer_cols]
            feedback_correct = cell(row, correct_expl_idx)
            feedback_incorrect = cell(row, incorrect_expl_idx)
            # Sanity: unexpected data beyond the feedback columns with no overflow.
            tail_start = (incorrect_expl_idx or 0) + 1
            if any(_nonempty(cell(row, i)) for i in range(tail_start, len(row))):
                warnings.append(
                    f"Q{n}: unexpected data in unlabeled columns despite "
                    f"max CorrectAns index <= 5 (kept in row_raw)."
                )

        options = [o for o in options if _nonempty(o)]
        accepted = [
            cell(row, answer_start + (i - 1))
            for i in correct_indices
            if cell(row, answer_start + (i - 1)) is not None
        ]

        rows.append(
            {
                "assessment_id": assessment_id or "",
                "assessment_title": title,
                "course_id": course_id or "",
                "unit_uid": "",
                "blockId": "",
                "group": cell(row, group_idx) if group_idx is not None else "",
                "question_number": n,
                "blockType": block_type,
                "description": cell(row, name_to_idx["Question"]),
                "options_raw": options,
                "configured_correct_answer_raw": correct_raw,
                "configured_accepted_answers_raw": accepted,
                "configured_score": "",
                "feedback_correct": feedback_correct,
                "feedback_incorrect": feedback_incorrect,
                "settings_raw": "",
                "row_raw": raw,
                "source": source,
                "extraction_timestamp": extraction_timestamp,
            }
        )

    stats = {
        "questions_count": len(rows),
        "type_breakdown": type_counts,
        "overflow_questions": overflow_count,
        "warnings": warnings,
        "blank_fields_note": (
            "assessment_id, course_id, unit_uid, blockId, configured_score and "
            "settings_raw are NOT present in the UI export. Pass --assessment-id / "
            "--course-id to stamp the first two; the rest stay blank by design."
        ),
    }
    return rows, raw_rows, stats
