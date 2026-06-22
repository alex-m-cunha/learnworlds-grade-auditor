#!/usr/bin/env python3
"""Generate a human-readable audit review sheet (XLSX) from reconciliation data.

Produces <run-dir>/revisao_auditoria.xlsx with:
  - Grade summary table at the top (one row per student, with live formulas)
  - Full detail table below (one row per submitted answer, all 810 rows)

The reviewer edits only the "Pontos Corrigidos" column; grade formulas update automatically.

Usage:
    python tools/generate_review_sheet.py --run-dir "output/pggf2/uc1/.../2026-06-20_..."
"""

from __future__ import annotations

import argparse
import csv
import sys
from datetime import datetime
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment, Border, Font, GradientFill, PatternFill, Side
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("openpyxl not installed — run: pip install openpyxl>=3.1", file=sys.stderr)
    sys.exit(1)


# ---------------------------------------------------------------------------
# Colours
# ---------------------------------------------------------------------------
CLR_HEADER_BG   = "1F3864"   # dark navy
CLR_HEADER_FG   = "FFFFFF"
CLR_SUMMARY_BG  = "2F5496"   # medium blue
CLR_SUMMARY_FG  = "FFFFFF"
CLR_RED_ROW     = "FCE4D6"   # light salmon — urgent
CLR_YLW_ROW     = "FFF2CC"   # light yellow — review
CLR_ORIG_BG     = "F2F2F2"   # light grey — read-only
CLR_CORR_BG     = "FFFF99"   # yellow — editable
CLR_MAX_BG      = "F2F2F2"   # light grey — read-only
CLR_NOTES_BG    = "DAEEF3"   # light blue — editable
CLR_ALT_ROW     = "F9F9F9"   # subtle alternating row tint
CLR_DELTA_POS   = "C6EFCE"   # green — grade improved
CLR_DELTA_NEG   = "FFC7CE"   # red — grade dropped

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=10) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _align(wrap=False, h="left", v="center") -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ---------------------------------------------------------------------------
# Block type → Portuguese
# ---------------------------------------------------------------------------
BT_LABELS = {
    "single": "Escolha múltipla",
    "tmc": "Escolha múltipla",
    "mcma": "Escolha múltipla (várias)",
    "tmcma": "Escolha múltipla (várias)",
    "tf": "Verdadeiro / Falso",
    "ttf": "Verdadeiro / Falso",
    "fillintheblankblock": "Preenchimento de lacunas",
    "dropdown": "Dropdown",
    "td": "Dropdown",
    "shorttext": "Texto curto",
    "tst": "Texto curto",
    "match": "Correspondência",
}


def _bt_label(bt: str) -> str:
    return BT_LABELS.get((bt or "").lower(), bt or "—")


# ---------------------------------------------------------------------------
# Flag → human-readable situation + priority
# ---------------------------------------------------------------------------
def _situacao(row: dict, inferred_confidence: str = "") -> tuple[str, str]:
    """Return (priority_emoji, situacao_text).

    inferred_confidence: confidence level looked up from inferred_answer_key.csv
    for rows where verifiable=="inferred". High confidence rows are not highlighted.
    """
    flag = (row.get("flag") or "").strip()
    verifiable = (row.get("verifiable") or "").strip()
    is_correct = (row.get("is_correct") or "").strip().lower()

    if flag == "answer_accepted_but_zero":
        return "🔴", "Resposta certa — 0 pontos"
    if flag == "answer_not_accepted_but_full":
        return "🔴", "Resposta não aceite mas pontuação máxima"
    if flag == "answer_correct_per_doc_but_zero":
        return "🔴", "Resposta correcta (gabarito docente) — 0 pontos"
    if flag == "fill_in_blank_over_answered":
        return "🔴", "Texto a mais — verificar se merece crédito"
    if verifiable == "inferred":
        if inferred_confidence == "high":
            return "", ""  # high confidence — no highlight needed
        if is_correct == "true":
            return "🟡", "Inferida — parecer correcto (confirmar)"
        if is_correct == "false":
            return "🟡", "Inferida — parecer incorrecto (confirmar)"
        return "🟡", "Inferida — confirmar"
    if verifiable == "inferred_low_confidence":
        return "🟡", "Inferida — baixa confiança (rever)"
    return "", ""


def _pts(value) -> float:
    """Parse a point value to a clean 2-decimal float.

    Converts via Decimal to avoid IEEE 754 imprecision inherited from the CSV
    (e.g. '19.270000000000000006' → 19.27 exactly).
    """
    try:
        d = Decimal(str(value or "0").strip()).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        return float(d)
    except (InvalidOperation, ValueError):
        return 0.0


def _read_csv(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with open(path, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


# ---------------------------------------------------------------------------
# Main generator
# ---------------------------------------------------------------------------
def generate(run_dir: Path, output_path: Path) -> None:
    report_rows = _read_csv(run_dir / "reconcile" / "reconciliation_report" / "reconciliation_report.csv")
    grade_rows  = _read_csv(run_dir / "reconcile" / "grade_reconciliation" / "grade_reconciliation.csv")

    if not report_rows:
        print("ERRO: reconciliation_report.csv não encontrado ou vazio.", file=sys.stderr)
        sys.exit(1)

    # Build inferred confidence lookup: description_key → confidence
    # Used to suppress yellow highlight for high-confidence inferred rows.
    _inferred_conf: dict[str, str] = {}
    _inferred_csv = run_dir / "answer_key" / "inferred_answer_key.csv"
    if _inferred_csv.exists():
        for r in _read_csv(_inferred_csv):
            _key = (r.get("question_text") or "").strip().lower()
            if _key:
                _inferred_conf[_key] = r.get("confidence", "")

    def _inferred_confidence_for(row: dict) -> str:
        desc = (row.get("description") or "").strip().lower()
        return _inferred_conf.get(desc, "")

    # First pass: identify question descriptions that have at least one 🔴 row.
    # All other rows for these questions will be highlighted yellow.
    flagged_descriptions: set[str] = set()
    for row in report_rows:
        pri, _ = _situacao(row, _inferred_confidence_for(row))
        if pri == "🔴":
            desc = (row.get("description") or "").strip()
            if desc:
                flagged_descriptions.add(desc)

    # Derive label/program from path
    parts = run_dir.parts
    try:
        label   = parts[-2].replace("-", " ").title()
        program = parts[-3].upper()
    except IndexError:
        label, program = "—", "—"

    # Build student grade index
    grade_index: dict[str, dict] = {}
    for g in grade_rows:
        grade_index[g["email"]] = g

    # Sort: red rows first, then yellow (inferred medium/low + question-flagged), then normal
    def _sort_key(r: dict) -> tuple:
        pri, _ = _situacao(r, _inferred_confidence_for(r))
        desc = (r.get("description") or "").strip()
        is_flagged_q = desc in flagged_descriptions
        if pri == "🔴":
            pri_order = 0
        elif pri == "🟡" or is_flagged_q:
            pri_order = 1
        else:
            pri_order = 2
        try:
            qn = int(r.get("question_number") or 999)
        except ValueError:
            qn = 999
        return (pri_order, (r.get("username") or "").lower(), qn)

    report_rows.sort(key=_sort_key)

    wb = Workbook()
    ws = wb.active
    ws.title = "Revisão de Auditoria"

    # -----------------------------------------------------------------------
    # Row 1 — Title
    # -----------------------------------------------------------------------
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = f"Revisão de Auditoria — {label} ({program})"
    title_cell.font = _font(bold=True, color=CLR_HEADER_FG, size=13)
    title_cell.fill = _fill(CLR_HEADER_BG)
    title_cell.alignment = _align(h="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:L2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Gerado em {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Fonte: reconciliation_report.csv"
    sub_cell.font = _font(color="888888", size=9)
    sub_cell.fill = _fill("EBF3FB")
    sub_cell.alignment = _align(h="center")
    ws.row_dimensions[2].height = 14

    # -----------------------------------------------------------------------
    # Rows 4+ — Grade summary
    # -----------------------------------------------------------------------
    SUMMARY_START = 4

    # We need to know where the detail table starts to write SUMIF formulas.
    # Detail table header = SUMMARY_START + 1 (students) + 2 (blank + label)
    n_students = len(grade_rows) if grade_rows else len({r["email"] for r in report_rows})
    DETAIL_HEADER_ROW = SUMMARY_START + n_students + 3  # +1 blank, +1 header label, +1 header row
    DETAIL_DATA_START = DETAIL_HEADER_ROW + 1
    DETAIL_DATA_END   = DETAIL_DATA_START + len(report_rows) - 1

    # Column letters for detail table (defined below, referenced here for formulas)
    # A=Prioridade B=Situação C=Aluno D=Email E=Nº F=Tipo G=Enunciado
    # H=Resposta I=Pontos Originais J=Pontos Corrigidos K=Pontuação Máx L=Notas Revisor
    COL_EMAIL_DETAIL   = "D"
    COL_ORIG_DETAIL    = "I"
    COL_CORR_DETAIL    = "J"
    COL_MAX_DETAIL     = "K"

    # Summary header
    summ_hdr = ws[f"A{SUMMARY_START}"]
    ws.merge_cells(f"A{SUMMARY_START}:L{SUMMARY_START}")
    summ_hdr.value = "Recálculo de Notas"
    summ_hdr.font = _font(bold=True, color=CLR_SUMMARY_FG, size=11)
    summ_hdr.fill = _fill(CLR_SUMMARY_BG)
    summ_hdr.alignment = _align(h="center")
    ws.row_dimensions[SUMMARY_START].height = 18

    # Summary column headers
    SUMMARY_COLS = ["Aluno", "Email", "Nota LW (%)", "Pontos Originais", "Pontos Máx",
                    "Pontos Corrigidos", "Nova Nota (%)", "Δ Nota"]
    SUMMARY_COL_WIDTHS = [24, 30, 14, 18, 14, 18, 16, 10]
    hdr_row = SUMMARY_START + 1
    for ci, (col_name, col_w) in enumerate(zip(SUMMARY_COLS, SUMMARY_COL_WIDTHS), start=1):
        col_letter = get_column_letter(ci)
        cell = ws[f"{col_letter}{hdr_row}"]
        cell.value = col_name
        cell.font = _font(bold=True, color="FFFFFF", size=10)
        cell.fill = _fill("2E75B6")
        cell.alignment = _align(h="center")
        cell.border = BORDER
        ws.column_dimensions[col_letter].width = col_w
    ws.row_dimensions[hdr_row].height = 16

    # Summary data rows — one per student sorted alphabetically
    students_sorted = sorted(grade_rows, key=lambda g: (g.get("username") or "").lower()) if grade_rows else []
    if not students_sorted:
        emails = sorted({r["email"] for r in report_rows})
        students_sorted = [{"email": e, "username": e, "official_grade": "", "sum_points": "", "sum_max": ""} for e in emails]

    summary_email_to_row: dict[str, int] = {}
    for si, g in enumerate(students_sorted):
        sr = SUMMARY_START + 2 + si
        summary_email_to_row[g["email"]] = sr
        email = g["email"]

        orig_pts = _pts(g.get("sum_points"))
        max_pts  = _pts(g.get("sum_max"))
        lw_grade  = g.get("official_grade", "")

        # SUMIF formulas referencing detail table
        corr_formula  = (f"=SUMIF(${COL_EMAIL_DETAIL}${DETAIL_DATA_START}:${COL_EMAIL_DETAIL}${DETAIL_DATA_END},"
                         f'B{sr},'
                         f"${COL_CORR_DETAIL}${DETAIL_DATA_START}:${COL_CORR_DETAIL}${DETAIL_DATA_END})")
        nova_nota_formula = f"=IF(E{sr}>0,F{sr}/E{sr}*100,\"\")"
        delta_formula     = f"=IF(AND(G{sr}<>\"\",C{sr}<>\"\"),G{sr}-C{sr},\"\")"

        row_data = [
            g.get("username", email),
            email,
            lw_grade,
            orig_pts,
            max_pts,
            corr_formula,
            nova_nota_formula,
            delta_formula,
        ]

        for ci, val in enumerate(row_data, start=1):
            col_letter = get_column_letter(ci)
            cell = ws[f"{col_letter}{sr}"]
            cell.value = val
            cell.border = BORDER
            cell.alignment = _align(h="center" if ci > 2 else "left")
            cell.font = _font(size=10)

            if ci in (4, 5):   # Pontos Originais, Pontos Máx
                cell.number_format = "0.00"
            if ci == 6:        # Pontos Corrigidos (fórmula)
                cell.font = _font(bold=True, size=10)
                cell.number_format = "0.00"
            if ci == 7:        # Nova Nota %
                cell.font = _font(bold=True, size=10)
                cell.number_format = "0.0"
            if ci == 8:        # Δ Nota
                cell.number_format = "+0.0;-0.0;0.0"
                cell.font = _font(bold=True, size=10)

        ws.row_dimensions[sr].height = 15

    # Freeze panes just below summary
    ws.freeze_panes = f"A{DETAIL_HEADER_ROW}"

    # -----------------------------------------------------------------------
    # Blank separator
    # -----------------------------------------------------------------------
    sep_row = DETAIL_HEADER_ROW - 2
    ws.merge_cells(f"A{sep_row}:L{sep_row}")
    ws[f"A{sep_row}"].value = "Detalhe das submissões — editar coluna «Pontos Corrigidos» para ajustar pontuação"
    ws[f"A{sep_row}"].font = _font(bold=True, color=CLR_SUMMARY_FG, size=10)
    ws[f"A{sep_row}"].fill = _fill(CLR_SUMMARY_BG)
    ws[f"A{sep_row}"].alignment = _align(h="center")
    ws.row_dimensions[sep_row].height = 16

    # -----------------------------------------------------------------------
    # Detail table header
    # -----------------------------------------------------------------------
    DETAIL_COLS = [
        ("Pri.",        6),
        ("Situação",   30),
        ("Aluno",      22),
        ("Email",      28),
        ("Nº",          6),
        ("Tipo",       20),
        ("Enunciado",  52),
        ("Resposta Submetida", 40),
        ("Pontos Originais",   16),
        ("Pontos Corrigidos",  16),
        ("Pontuação Máx",      14),
        ("Notas Revisor",      32),
    ]

    for ci, (col_name, col_w) in enumerate(DETAIL_COLS, start=1):
        col_letter = get_column_letter(ci)
        cell = ws[f"{col_letter}{DETAIL_HEADER_ROW}"]
        cell.value = col_name
        cell.font = _font(bold=True, color=CLR_HEADER_FG, size=10)
        cell.fill = _fill(CLR_HEADER_BG)
        cell.alignment = _align(h="center")
        cell.border = BORDER
        ws.column_dimensions[col_letter].width = col_w
    ws.row_dimensions[DETAIL_HEADER_ROW].height = 16

    # -----------------------------------------------------------------------
    # Detail data rows
    # -----------------------------------------------------------------------
    for ri, row in enumerate(report_rows):
        dr = DETAIL_DATA_START + ri
        priority, situacao = _situacao(row, _inferred_confidence_for(row))

        pts_orig = _pts(row.get("points"))
        pts_max  = _pts(row.get("max_points"))

        qn = row.get("question_number") or "—"
        desc = (row.get("description") or "").strip()
        cells_data = [
            priority,
            situacao,
            row.get("username") or "",
            row.get("email") or "",
            qn,
            _bt_label(row.get("blockType") or ""),
            desc[:300],
            (row.get("submitted_answer") or "")[:500],
            pts_orig,   # Pontos Originais
            pts_orig,   # Pontos Corrigidos (pre-filled = original)
            pts_max,    # Pontuação Máx
            "",         # Notas Revisor
        ]

        # Colour logic:
        # • ALL rows of a question that has any 🔴 case → yellow row background
        # • Within those rows, the 🔴 case itself gets specific cells in red:
        #     col 1 (Pri.), col 2 (Situação), col 8 (Resposta), col 9 (Pontos Originais)
        # • Medium/low inferred rows not part of a flagged question → yellow row
        # • Everything else → subtle alternating grey
        is_flagged_question = desc in flagged_descriptions
        is_red_case = priority == "🔴"
        flag = (row.get("flag") or "").strip()
        zero_score_flag = flag in ("answer_accepted_but_zero", "answer_correct_per_doc_but_zero")

        if is_flagged_question or priority == "🟡":
            row_fill = _fill(CLR_YLW_ROW)
        else:
            row_fill = _fill(CLR_ALT_ROW) if ri % 2 == 0 else None

        # Columns to paint red for the specific detected case
        red_cells: set[int] = set()
        if is_red_case:
            red_cells = {1, 2, 8}          # Pri., Situação, Resposta
            if zero_score_flag:
                red_cells.add(9)           # also Pontos Originais for zero-score issues

        for ci, val in enumerate(cells_data, start=1):
            col_letter = get_column_letter(ci)
            cell = ws[f"{col_letter}{dr}"]
            cell.value = val
            cell.font = _font(size=9)
            cell.border = BORDER

            # Alignment
            if ci in (1, 5, 9, 10, 11):
                cell.alignment = _align(h="center")
            elif ci in (7, 8, 12):
                cell.alignment = _align(wrap=True)
            else:
                cell.alignment = _align()

            # Fill priority: red cell > special column > row fill
            if ci in red_cells:
                cell.fill = _fill(CLR_RED_ROW)
                cell.font = _font(bold=True, size=9)
            elif ci == 9:   # Pontos Originais — read-only
                cell.fill = _fill(CLR_ORIG_BG)
                cell.number_format = "0.00"
            elif ci == 10:  # Pontos Corrigidos — editable
                cell.fill = _fill(CLR_CORR_BG)
                cell.font = _font(bold=True, size=9)
                cell.number_format = "0.00"
            elif ci == 11:  # Pontuação Máx — read-only
                cell.fill = _fill(CLR_MAX_BG)
                cell.number_format = "0.00"
            elif ci == 12:  # Notas Revisor — editable
                cell.fill = _fill(CLR_NOTES_BG)
            elif row_fill:
                cell.fill = row_fill

        ws.row_dimensions[dr].height = 30 if situacao else 20

    # -----------------------------------------------------------------------
    # Conditional formatting for Δ Nota in summary (green/red)
    # -----------------------------------------------------------------------
    from openpyxl.formatting.rule import CellIsRule
    delta_col = "H"
    delta_range = f"{delta_col}{SUMMARY_START + 2}:{delta_col}{SUMMARY_START + 1 + len(students_sorted)}"
    ws.conditional_formatting.add(
        delta_range,
        CellIsRule(operator="greaterThan", formula=["0"], fill=_fill(CLR_DELTA_POS))
    )
    ws.conditional_formatting.add(
        delta_range,
        CellIsRule(operator="lessThan", formula=["0"], fill=_fill(CLR_DELTA_NEG))
    )

    # -----------------------------------------------------------------------
    # Auto-filter on detail table — pre-applied to show only 🔴 (red-cell) rows
    # -----------------------------------------------------------------------
    from openpyxl.worksheet.filters import FilterColumn, Filters

    ws.auto_filter.ref = f"A{DETAIL_HEADER_ROW}:{get_column_letter(len(DETAIL_COLS))}{DETAIL_DATA_END}"

    # Set filter definition on column A (colId=0 = first column of the filter range)
    fc = FilterColumn(colId=0)
    fc.filters = Filters(filter=["🔴"])
    ws.auto_filter.filterColumn = [fc]

    # Hide rows that don't match the filter so Excel shows the view immediately on open
    for ri, row in enumerate(report_rows):
        dr = DETAIL_DATA_START + ri
        pri, _ = _situacao(row, _inferred_confidence_for(row))
        if pri != "🔴":
            ws.row_dimensions[dr].hidden = True

    # -----------------------------------------------------------------------
    # Save
    # -----------------------------------------------------------------------
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Folha de revisão escrita em: {output_path}")
    print(f"  {len(report_rows)} linhas de detalhe  |  {len(students_sorted)} alunos")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def _parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Generate audit review XLSX from reconciliation data.")
    p.add_argument("--run-dir", required=True, help="Path to the timestamped run folder")
    p.add_argument("--output", default="", help="Output .xlsx path (default: <run-dir>/revisao_auditoria.xlsx)")
    return p.parse_args()


def main() -> None:
    args = _parse_args()
    run_dir = Path(args.run_dir)
    if not run_dir.is_dir():
        print(f"ERRO: pasta não encontrada: {run_dir}", file=sys.stderr)
        sys.exit(1)
    out = Path(args.output) if args.output else run_dir / "revisao_auditoria.xlsx"
    generate(run_dir, out)


if __name__ == "__main__":
    main()
